# -*- coding: utf-8 -*-
"""
Módulo de tratamento e análise de dados — Sistema de Ajustes & Ocorrências.

Responsável por:
- Carregar a planilha enviada pelo usuário (xlsx/xls/csv)
- Normalizar nomes de colunas (robusto a variações de cabeçalho)
- Classificar automaticamente a causa de cada ocorrência a partir do texto
  da observação (regras de negócio baseadas em palavras-chave)
- Calcular os indicadores (KPIs) e séries usadas no dashboard
"""

from __future__ import annotations

import io
import re
import unicodedata
from dataclasses import dataclass, field

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)


# --------------------------------------------------------------------------- #
# Normalização de texto
# --------------------------------------------------------------------------- #

def _strip_accents(text: str) -> str:
    text = unicodedata.normalize("NFD", str(text))
    return "".join(c for c in text if unicodedata.category(c) != "Mn")


def _clean(text: str) -> str:
    return _strip_accents(str(text)).strip().lower()


# --------------------------------------------------------------------------- #
# Mapeamento flexível de colunas de entrada
# --------------------------------------------------------------------------- #

COLUMN_ALIASES = {
    "OM": ["om", "ordem", "ordem de manutencao", "n ordem", "numero", "id"],
    "OFICINA": ["oficina", "fornecedor", "parceiro", "unidade", "fabrica"],
    "DATA": ["data", "data ocorrencia", "data ajuste", "dt"],
    "DESCRICAO": [
        "descricao obs",
        "descricao",
        "obs",
        "observacao",
        "motivo",
        "causa",
        "detalhe",
    ],
}


def _match_column(columns: list[str], aliases: list[str]) -> str | None:
    cleaned = {c: _clean(c) for c in columns}
    for col, c in cleaned.items():
        if c in aliases:
            return col
    for col, c in cleaned.items():
        if any(alias in c for alias in aliases):
            return col
    return None


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Renomeia as colunas da planilha original para os nomes canônicos
    usados internamente: OM, OFICINA, DATA, DESCRICAO."""
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    rename_map = {}
    for canonical, aliases in COLUMN_ALIASES.items():
        found = _match_column(list(df.columns), aliases)
        if found:
            rename_map[found] = canonical
    df = df.rename(columns=rename_map)

    for required in ["OFICINA", "DATA", "DESCRICAO"]:
        if required not in df.columns:
            raise ValueError(
                f"Não foi possível localizar a coluna correspondente a "
                f"'{required}' na planilha enviada. Verifique o cabeçalho do arquivo."
            )
    if "OM" not in df.columns:
        df["OM"] = range(1, len(df) + 1)

    return df[["OM", "OFICINA", "DATA", "DESCRICAO"]]


# --------------------------------------------------------------------------- #
# Classificação de causas (regras de negócio)
# --------------------------------------------------------------------------- #

# Ordem importa: a primeira regra que casar define a causa.
CAUSE_RULES: list[tuple[str, list[str]]] = [
    ("Descontos", ["desconto"]),
    ("Balanceamento", ["balanceamento"]),
    ("Reembolso", ["reembolso"]),
    ("Arredondamento", ["arredond"]),
    ("Tempo Divergente", ["tempo divergente", "tempo/preco", "tempo"]),
    ("Valor Unitário", ["unit", "valor errado", "preco", "valor"]),
    ("Erro de Cálculo", ["calculo"]),
    (
        "Integração/Sistema (GERFAC)",
        ["integra", "conector", "gerfac", "critica", "suporte", "falta"],
    ),
]

# Causas consideradas de origem financeira (pedido explícito da gestão).
FINANCIAL_CAUSES = ["Descontos", "Balanceamento", "Valor Unitário", "Reembolso"]


def classify_cause(description: str) -> str:
    text = _clean(description)
    for label, keywords in CAUSE_RULES:
        if any(keyword in text for keyword in keywords):
            return label
    return "Outros"


# --------------------------------------------------------------------------- #
# Carregamento principal
# --------------------------------------------------------------------------- #

def load_dataframe(uploaded_file) -> pd.DataFrame:
    """Lê o arquivo enviado (xlsx/xls/csv) em um DataFrame bruto."""
    name = getattr(uploaded_file, "name", "") or ""
    if name.lower().endswith(".csv"):
        raw = pd.read_csv(uploaded_file)
    else:
        raw = pd.read_excel(uploaded_file)
    return raw


def process_data(raw_df: pd.DataFrame) -> pd.DataFrame:
    """Pipeline completo de tratamento: normaliza colunas, tipa datas,
    remove lixo e classifica a causa de cada ocorrência."""
    df = normalize_columns(raw_df)

    # Tratamento de datas: aceita tanto datas já tipadas quanto texto.
    df["DATA"] = pd.to_datetime(df["DATA"], dayfirst=True, errors="coerce")

    # Linhas sem oficina ou sem descrição não são ocorrências válidas.
    df["OFICINA"] = df["OFICINA"].astype(str).str.strip()
    df["DESCRICAO"] = df["DESCRICAO"].astype(str).str.strip()
    df = df[(df["OFICINA"] != "") & (df["OFICINA"].str.lower() != "nan")]
    df = df[df["DATA"].notna()]

    df["CAUSA"] = df["DESCRICAO"].apply(classify_cause)
    df["DIA"] = df["DATA"].dt.date

    df = df.reset_index(drop=True)
    return df


# --------------------------------------------------------------------------- #
# KPIs e agregações usadas pelo dashboard
# --------------------------------------------------------------------------- #

@dataclass
class DashboardMetrics:
    total_ocorrencias: int = 0
    media_diaria: float = 0.0
    causa_principal: str = "-"
    causa_principal_qtd: int = 0
    causa_principal_pct: float = 0.0
    segunda_causa: str = "-"
    segunda_causa_qtd: int = 0
    segunda_causa_pct: float = 0.0
    pct_financeiro: float = 0.0
    qtd_oficinas_reincidentes: int = 0
    qtd_oficinas_total: int = 0
    periodo_inicio: str = "-"
    periodo_fim: str = "-"
    causas_resumo: pd.DataFrame = field(default_factory=pd.DataFrame)
    oficinas_resumo: pd.DataFrame = field(default_factory=pd.DataFrame)
    serie_diaria: pd.DataFrame = field(default_factory=pd.DataFrame)
    # Pareto de oficinas
    pareto_oficinas: pd.DataFrame = field(default_factory=pd.DataFrame)
    pareto_80pct_count: int = 0
    pareto_80pct_pct_oficinas: float = 0.0
    # Qualidade de classificação — causa "Outros"
    outros_qtd: int = 0
    outros_pct: float = 0.0
    outros_top_descricoes: pd.DataFrame = field(default_factory=pd.DataFrame)


def build_metrics(df: pd.DataFrame) -> DashboardMetrics:
    m = DashboardMetrics()
    total = len(df)
    m.total_ocorrencias = total
    if total == 0:
        return m

    dias_distintos = df["DIA"].nunique() or 1
    m.media_diaria = round(total / dias_distintos, 1)

    causas = (
        df["CAUSA"]
        .value_counts()
        .rename_axis("CAUSA")
        .reset_index(name="QTD")
    )
    causas["PERCENTUAL"] = (causas["QTD"] / total * 100).round(1)
    m.causas_resumo = causas

    if len(causas) > 0:
        m.causa_principal = causas.iloc[0]["CAUSA"]
        m.causa_principal_qtd = int(causas.iloc[0]["QTD"])
        m.causa_principal_pct = float(causas.iloc[0]["PERCENTUAL"])
    if len(causas) > 1:
        m.segunda_causa = causas.iloc[1]["CAUSA"]
        m.segunda_causa_qtd = int(causas.iloc[1]["QTD"])
        m.segunda_causa_pct = float(causas.iloc[1]["PERCENTUAL"])

    financeiro_qtd = df[df["CAUSA"].isin(FINANCIAL_CAUSES)].shape[0]
    m.pct_financeiro = round(financeiro_qtd / total * 100, 1)

    oficinas_count = (
        df["OFICINA"]
        .value_counts()
        .rename_axis("OFICINA")
        .reset_index(name="QTD")
        .sort_values("QTD", ascending=False)
    )
    m.oficinas_resumo = oficinas_count
    m.qtd_oficinas_total = oficinas_count.shape[0]
    m.qtd_oficinas_reincidentes = int((oficinas_count["QTD"] >= 2).sum())

    # Pareto de oficinas — % acumulado por volume (decrescente)
    pareto = oficinas_count.copy().reset_index(drop=True)
    pareto["PCT"] = (pareto["QTD"] / total * 100).round(2)
    cumsum_raw = pareto["PCT"].cumsum()
    pareto["PCT_ACUM"] = cumsum_raw.round(1)
    m.pareto_oficinas = pareto
    if (cumsum_raw >= 80.0).any():
        m.pareto_80pct_count = int((cumsum_raw >= 80.0).idxmax()) + 1
    else:
        m.pareto_80pct_count = len(pareto)
    m.pareto_80pct_pct_oficinas = (
        round(m.pareto_80pct_count / m.qtd_oficinas_total * 100, 1)
        if m.qtd_oficinas_total else 0.0
    )

    # Qualidade de classificação — causa "Outros"
    outros_mask = df["CAUSA"] == "Outros"
    m.outros_qtd = int(outros_mask.sum())
    m.outros_pct = round(m.outros_qtd / total * 100, 1)
    if m.outros_qtd > 0:
        m.outros_top_descricoes = (
            df.loc[outros_mask, "DESCRICAO"]
            .value_counts()
            .head(10)
            .rename_axis("DESCRICAO")
            .reset_index(name="QTD")
        )

    serie = (
        df.groupby("DIA")
        .size()
        .rename("QTD")
        .reset_index()
        .sort_values("DIA")
    )
    serie["MEDIA_MOVEL"] = serie["QTD"].rolling(window=3, min_periods=1).mean().round(1)
    m.serie_diaria = serie

    m.periodo_inicio = pd.to_datetime(df["DATA"].min()).strftime("%d/%m/%Y")
    m.periodo_fim = pd.to_datetime(df["DATA"].max()).strftime("%d/%m/%Y")

    return m


# --------------------------------------------------------------------------- #
# Exportação — planilha executiva (.xlsx)
# --------------------------------------------------------------------------- #

def export_excel(df: pd.DataFrame, periodo_inicio: str, periodo_fim: str) -> bytes:
    """Gera um arquivo Excel executivo com os registros de df (já filtrado).

    Colunas exportadas: OM, OFICINA, DATA, CAUSA, DESCRICAO — idênticas à
    tabela exibida no dashboard. Layout: cabeçalho escuro com texto aqua,
    zebra-striping suave, bordas finas, larguras ajustadas e aba de metadados.
    """
    # ── paleta ──────────────────────────────────────────────────────────────
    C_HEADER_BG  = "070B10"
    C_HEADER_FG  = "1FE7B8"
    C_ROW_ALT    = "0E151B"
    C_ROW_EVEN   = "0B1117"
    C_BORDER     = "1FE7B822"
    C_TEXT       = "E9F5F1"
    C_TEXT_DIM   = "7F93A0"
    C_ACCENT     = "1FE7B8"

    thin = Side(style="thin", color="1A2E26")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_font  = Font(name="Calibri", bold=True, size=10, color=C_HEADER_FG)
    header_fill  = PatternFill("solid", fgColor=C_HEADER_BG)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=False)

    data_font      = Font(name="Calibri", size=10, color=C_TEXT)
    data_font_dim  = Font(name="Calibri", size=10, color=C_TEXT_DIM)
    data_align_l   = Alignment(horizontal="left",   vertical="center", wrap_text=False)
    data_align_c   = Alignment(horizontal="center", vertical="center", wrap_text=False)
    data_align_wrap = Alignment(horizontal="left",  vertical="center", wrap_text=True)

    fill_even = PatternFill("solid", fgColor=C_ROW_EVEN)
    fill_odd  = PatternFill("solid", fgColor=C_ROW_ALT)

    wb = Workbook()
    ws = wb.active
    ws.title = "Dados Tratados"

    # ── cabeçalho da planilha (título executivo) ─────────────────────────────
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = "AJUSTES & OCORRÊNCIAS — RESUMO EXECUTIVO"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color=C_ACCENT)
    title_cell.fill = PatternFill("solid", fgColor=C_HEADER_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:E2")
    sub_cell = ws["A2"]
    sub_cell.value = f"Período: {periodo_inicio} a {periodo_fim}  |  Total: {len(df):,} ocorrências"
    sub_cell.font = Font(name="Calibri", size=10, color=C_TEXT_DIM)
    sub_cell.fill = PatternFill("solid", fgColor=C_HEADER_BG)
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    ws.merge_cells("A3:E3")
    ws["A3"].fill = PatternFill("solid", fgColor=C_HEADER_BG)
    ws.row_dimensions[3].height = 6

    # ── linha de cabeçalho das colunas (linha 4) ────────────────────────────
    COLUMNS = ["OM", "OFICINA", "DATA", "CAUSA", "DESCRIÇÃO"]
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=4, column=col_idx, value=col_name)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = header_align
        cell.border = border
    ws.row_dimensions[4].height = 22

    # ── dados (a partir da linha 5) ─────────────────────────────────────────
    export_df = df[["OM", "OFICINA", "DATA", "CAUSA", "DESCRICAO"]].copy()
    export_df = export_df.sort_values("DATA", ascending=False).reset_index(drop=True)
    export_df["DATA"] = pd.to_datetime(export_df["DATA"]).dt.strftime("%d/%m/%Y")

    for row_idx, (_, row) in enumerate(export_df.iterrows(), start=5):
        fill = fill_even if row_idx % 2 == 0 else fill_odd
        values = [row["OM"], row["OFICINA"], row["DATA"], row["CAUSA"], row["DESCRICAO"]]
        aligns = [data_align_c, data_align_l, data_align_c, data_align_c, data_align_wrap]
        fonts  = [data_font_dim, data_font, data_font_dim, data_font, data_font_dim]

        for col_idx, (val, aln, fnt) in enumerate(zip(values, aligns, fonts), start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = fnt
            cell.fill      = fill
            cell.alignment = aln
            cell.border    = border
        ws.row_dimensions[row_idx].height = 16

    # ── largura das colunas ─────────────────────────────────────────────────
    col_widths = {"A": 14, "B": 30, "C": 14, "D": 22, "E": 60}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # ── congelar painel no cabeçalho de colunas ─────────────────────────────
    ws.freeze_panes = "A5"

    # ── auto-filtro nos cabeçalhos ───────────────────────────────────────────
    ws.auto_filter.ref = f"A4:E{4 + len(export_df)}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
